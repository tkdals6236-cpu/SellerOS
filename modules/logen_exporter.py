# ====================================
# SellerOS Logen Exporter
# ====================================
#
# 역할
# - match_only=True  : MATCH만 출력 (검수 결과)
# - match_only=False : 모든 주문 출력 (직접 입력)
#
# 작성 : 김상민
# 버전 : 0.0.4
# ====================================

from openpyxl import load_workbook
from config import LOGEN_TEMPLATE


def export_logen_excel(results, file_path, match_only=True):

    # ------------------------
    # 로젠 원본 양식 불러오기
    # ------------------------
    wb = load_workbook(LOGEN_TEMPLATE)
    ws = wb.active

    # 예제 데이터 삭제
    ws.delete_rows(1, 2)

    # ------------------------
    # 동일 주문자 묶기
    # ------------------------
    grouped = {}

    for row in results:

        # 검수 결과에서는 MATCH만 출력
        if match_only and row.get("status") != "MATCH":
            continue

        # 이름
        name = row.get("nickname") or row.get("depositor", "")

        key = (
            name,
            row.get("phone", ""),
            row.get("address", "")
        )

        if key not in grouped:

            new_row = row.copy()
            new_row["count"] = 1
            grouped[key] = new_row

        else:

            grouped[key]["count"] += 1

    # ------------------------
    # 로젠 양식 출력
    # ------------------------
    current_row = 1

    for row in grouped.values():

        name = row.get("nickname") or row.get("depositor", "")

        # 품명 표시
        if row["count"] == 1:
            product_name = row.get("product", "")
        else:
            product_name = f'{row.get("product", "")} 외 {row["count"] - 1}건'

        ws[f"A{current_row}"] = name
        ws[f"B{current_row}"] = ""
        ws[f"C{current_row}"] = row.get("address", "")
        ws[f"D{current_row}"] = ""
        ws[f"E{current_row}"] = row.get("phone", "")
        ws[f"F{current_row}"] = 1
        ws[f"G{current_row}"] = ""
        ws[f"H{current_row}"] = "010"
        ws[f"I{current_row}"] = product_name
        ws[f"J{current_row}"] = ""
        ws[f"K{current_row}"] = "친절 빠른배송 부탁드립니다."

        current_row += 1

    wb.save(file_path)