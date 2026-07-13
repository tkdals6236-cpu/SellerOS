from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def save_order_list_excel(orders, file_path):

    wb = Workbook()
    ws = wb.active

    ws.title = "주문목록"

    ws.append([
        "확인",
        "닉네임",
        "상품",
        "입금자",
        "연락처",
        "주소",
        "비고"
    ])

    header_fill = PatternFill(
        fill_type="solid",
        start_color="A9D18E"
    )

    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for order in orders:

        ws.append([
            "☐",
            order.get("nickname", ""),
            order.get("product", ""),
            order.get("depositor", ""),
            order.get("phone", ""),
            order.get("address", ""),
            ""
        ])

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 50
    ws.column_dimensions["G"].width = 18

    wb.save(file_path)