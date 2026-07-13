# ====================================
# SellerOS Validator Module
# ====================================
#
# 역할
# - 검수 결과를 엑셀로 저장한다.
# - 셀러가 검수하기 쉬운 형태로 출력한다.
#
# 작성 : 김상민
# 버전 : 0.0.3
# ====================================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def save_result_excel(results, file_path):

    wb = Workbook()
    ws = wb.active

    ws.title = "분석결과"

    # 헤더
    ws.append([
       "확인",
       "상태",
       "닉네임",
       "상품",
       "입금자",
       "입금금액",
       "연락처",
       "주소",
       "입금일시",
       "비고"
])

# ------------------------
# 상태 표시용
# ------------------------
    status_text = {
        "MATCH": "입금완료",
        "NO_PAYMENT": "미입금",
        "NO_ORDER" : "주문서없음"
}

# ------------------------
# 헤더 꾸미기
# ------------------------
    header_fill = PatternFill(
        fill_type="solid",
        start_color="A9D18E"
    )
    match_fill = PatternFill(
        fill_type="solid",
        start_color="D9EAD3"   # 연한 초록
    )
    no_payment_fill = PatternFill(
        fill_type="solid",
        start_color="F4CCCC"   # 연한 빨강
    )
    no_order_fill = PatternFill(
        fill_type="solid",
        start_color="FFF2CC"   #주황
    )
    
    
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ------------------------
    # 데이터
    # ------------------------
    for row in results:

        ws.append([
            "☐",
            status_text[row["status"]],
            row["nickname"],
            row["product"],
            row["depositor"],
            f'{row["amount"]:,}' if row["amount"] else "",
            row["phone"],
            row["address"],
            row["datetime"]
        ])
        current_row = ws.max_row

    # 정렬
        ws[f"A{current_row}"].alignment = Alignment(horizontal="center")
        ws[f"B{current_row}"].alignment = Alignment(horizontal="center")
        ws[f"C{current_row}"].alignment = Alignment(horizontal="center")
        ws[f"D{current_row}"].alignment = Alignment(horizontal="left")
        ws[f"E{current_row}"].alignment = Alignment(horizontal="center")
        ws[f"F{current_row}"].alignment = Alignment(horizontal="right")
        ws[f"G{current_row}"].alignment = Alignment(horizontal="center")
        ws[f"H{current_row}"].alignment = Alignment(horizontal="left")
        ws[f"I{current_row}"].alignment = Alignment(horizontal="center")
        ws[f"J{current_row}"].alignment = Alignment(horizontal="left")
        status_cell = ws[f"A{current_row}"]
    
        if row["status"] == "MATCH":
            status_cell.fill = match_fill

        elif row["status"] == "NO_PAYMENT":
            status_cell.fill = no_payment_fill

        elif row["status"] == "NO_ORDER":
            status_cell.fill = no_order_fill
            
    
    # 컬럼 너비
    ws.column_dimensions["A"].width = 8    # 확인
    ws.column_dimensions["B"].width = 12   # 상태
    ws.column_dimensions["C"].width = 18   # 닉네임
    ws.column_dimensions["D"].width = 35   # 상품
    ws.column_dimensions["E"].width = 18   # 입금자
    ws.column_dimensions["F"].width = 15   # 입금금액
    ws.column_dimensions["G"].width = 18   # 연락처
    ws.column_dimensions["H"].width = 55   # 주소
    ws.column_dimensions["I"].width = 22   # 입금일시
    ws.column_dimensions["J"].width = 20   # 비고
    wb.save(file_path)