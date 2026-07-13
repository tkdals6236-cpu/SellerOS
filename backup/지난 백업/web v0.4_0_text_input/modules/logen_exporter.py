# ====================================
# SellerOS Logen Exporter
# ====================================
#
# 역할
# - MATCH 데이터만
#   로젠 양식에 입력한다.
#
# 작성 : 김상민
# 버전 : 0.0.3
# ====================================

from openpyxl import load_workbook
from config import LOGEN_TEMPLATE


def export_logen_excel(results, file_path):

    # ------------------------
    # 로젠 원본 양식 불러오기
    # ------------------------
    wb = load_workbook(LOGEN_TEMPLATE)
    ws = wb.active

    # 예제 데이터 삭제
    ws.delete_rows(1, 2)

    # ------------------------
    # 동일 주문자 묶기
    # ------------------------
    grouped = {}

    for row in results:

        if row["status"] != "MATCH":
            continue

        key = (
            row["nickname"],
            row["phone"],
            row["address"]
        )

        if key not in grouped:

            new_row = row.copy()
            new_row["count"] = 1
            grouped[key] = new_row

        else:

            grouped[key]["count"] += 1

    # ------------------------
    # 로젠 양식 출력
    # ------------------------
    current_row = 1

    for row in grouped.values():

        # 품명 표시
        if row["count"] == 1:
            product_name = row["product"]
        else:
            product_name = f'{row["product"]} 외 {row["count"] - 1}건'

        ws[f"A{current_row}"] = row["nickname"]     # 이름
        ws[f"B{current_row}"] = ""                  # 공란
        ws[f"C{current_row}"] = row["address"]      # 주소
        ws[f"D{current_row}"] = ""                  # 전화번호
        ws[f"E{current_row}"] = row["phone"]        # 휴대폰
        ws[f"F{current_row}"] = 1                   # 수량
        ws[f"G{current_row}"] = ""                  # 운송료
        ws[f"H{current_row}"] = "010"               # 선불
        ws[f"I{current_row}"] = product_name        # 품명
        ws[f"J{current_row}"] = ""                  # 공란
        ws[f"K{current_row}"] = "친절 빠른배송 부탁드립니다."

        current_row += 1

    wb.save(file_path)