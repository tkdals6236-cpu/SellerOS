from flask import Flask, render_template, request, send_file, session, redirect
import os
import openpyxl
from werkzeug.utils import secure_filename
from modules.pipeline import run_pipeline
from modules.error_handler import SellerOSError
from modules.order_text_parser import load_order_text
from modules.order_parser import load_order_excel
from modules.order_exporter import save_order_list_excel
from datetime import datetime
from modules.group_preview import group_preview
import shutil
import time
from modules.preview_manager import get_preview
from flask import jsonify

ADMIN_ID = "admin"
ADMIN_PW = "1234!!"

app = Flask(__name__)
app.secret_key = "selleros_dev"

os.makedirs("uploads/orders", exist_ok=True)
os.makedirs("uploads/banks", exist_ok=True)
os.makedirs("uploads/files", exist_ok=True)
os.makedirs("output", exist_ok=True)



def clean_old_files():

    
    folders = [
        "uploads/orders",
        "uploads/banks",
        "output"
    ]

    now = time.time()

    for folder in folders:

        if not os.path.exists(folder):
            continue

        for filename in os.listdir(folder):

            file_path = os.path.join(folder, filename)

            try:

                if not os.path.isfile(file_path):
                    continue

                age = now - os.path.getmtime(file_path)

                # 임시 파일은 30분(1800초) 지난 파일 삭제
                if age > 1800:
                    
                    os.remove(file_path)

            except Exception as e:

                print(f"삭제 실패 : {file_path} ({e})")

# clean_temp_folders()

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/panel_4d9a1f")
def admin_login_page():
    return render_template("admin_login.html")

@app.route("/panel_4d9a1f/home")
def admin_home():

    if not session.get("admin"):
        return redirect("/")

    return render_template("admin_home.html")

@app.route("/logout")
def logout():

    session.pop("admin", None)

    return redirect("/")

@app.route("/file")
def file_manager():

    if not session.get("admin"):
        return redirect("/")

    folder = os.path.join("uploads", "files")

    files = []

    if os.path.exists(folder):

        for filename in os.listdir(folder):

            path = os.path.join(folder, filename)

            display_name = filename

            # 앞에 붙인 날짜 제거
            if len(filename) > 16 and filename[15] == "_":
                display_name = filename[16:]

            files.append({

                "name": filename,
                "display": display_name,

                "mtime": datetime.fromtimestamp(
                    os.path.getmtime(path)
                ).strftime("%Y-%m-%d %H:%M")

            })

        files.sort(
            key=lambda x: x["mtime"],
            reverse=True
        )

    return render_template(
        "file_manager.html",
        files=files
    )

@app.route("/order")
def order_manager():

    if not session.get("admin"):
        return redirect("/")

    return render_template("order_manager.html")


@app.route("/stock")
def stock_manager():

    if not session.get("admin"):
        return redirect("/")

    return render_template("stock_manager.html")


@app.route("/client")
def client_manager():

    if not session.get("admin"):
        return redirect("/")

    return render_template("client_manager.html")


@app.route("/settings")
def settings():

    if not session.get("admin"):
        return redirect("/")

    return render_template("settings.html")

@app.route("/admin_auth", methods=["POST"])
def admin_auth():

    admin_id = request.form.get("admin_id")
    admin_pw = request.form.get("admin_pw")

    if admin_id == ADMIN_ID and admin_pw == ADMIN_PW:

        session["admin"] = True

        return redirect("/file")

    return render_template(
        "admin_login.html",
        error="아이디 또는 비밀번호가 올바르지 않습니다."
    )

@app.route("/upload_file", methods=["POST"])
def upload_file():

    if not session.get("admin"):
        return redirect("/")

    file = request.files.get("file")

    if not file or file.filename == "":
        return redirect("/file")

    print("원본 파일명 :", file.filename)

    import os
    from datetime import datetime

    filename = os.path.basename(file.filename)

    # 같은 이름 방지
    filename = (
        datetime.now().strftime("%Y%m%d_%H%M%S_")
        + filename
    )

    save_path = os.path.join(
        "uploads",
        "files",
        filename
    )

    file.save(save_path)

    return redirect("/file")

@app.route("/uploads/files/<filename>")
def uploaded_file(filename):

    return send_file(
        os.path.join(
            "uploads",
            "files",
            filename
        )
    )

@app.route("/delete_file/<filename>")
def delete_file(filename):

    if not session.get("admin"):
        return redirect("/")

    # 핵심 엑셀 파일 삭제 방지
    if filename.endswith("_엑셀.xlsx") or filename == "엑셀.xlsx":
        return redirect("/file")

    file_path = os.path.join(
        "uploads",
        "files",
        filename
    )

    if os.path.exists(file_path):
        os.remove(file_path)

    return redirect("/file")

@app.route("/excel_preview/<path:filename>")
def excel_preview(filename):

    if not session.get("admin"):
        return {}

    import pandas as pd

    file_path = os.path.join("uploads", "files", filename)

    df = pd.read_excel(file_path)

    df = df.head(20)

    return df.to_html(index=False)

@app.route("/excel_data/<path:filename>")
def excel_data(filename):

    if not session.get("admin"):
        return jsonify({"error": "unauthorized"}), 401

    import openpyxl

    filename = os.path.basename(filename)

    file_path = os.path.join(
        "uploads",
        "files",
        filename
    )

    if not os.path.exists(file_path):
        return jsonify({
            "error": "파일을 찾을 수 없습니다."
        }), 404

    try:

        workbook = openpyxl.load_workbook(
            file_path,
            data_only=False
        )

        sheet = workbook.active

        cells = {}

        for row in range(1, sheet.max_row + 1):

            cells[row] = {}

            for col in range(1, sheet.max_column + 1):

                value = sheet.cell(
                    row=row,
                    column=col
                ).value

                if value is None:
                    value = ""

                cells[row][col] = str(value)

        return jsonify({

            "max_row": sheet.max_row,
            "max_col": sheet.max_column,
            "cells": cells

        })

    except Exception as e:

        return jsonify({
            "error": str(e)
        }), 500


@app.route("/save_excel", methods=["POST"])
def save_excel():

    if not session.get("admin"):
        return jsonify({
            "error": "unauthorized"
        }), 401

    try:

        data = request.get_json() or {}

        filename = os.path.basename(
            data.get("filename", "")
        )

        rows = data.get("data", [])

        if not filename:
            return jsonify({
                "error": "파일명이 없습니다."
            }), 400

        file_path = os.path.join(
            "uploads",
            "files",
            filename
        )

        if not os.path.exists(file_path):
            return jsonify({
                "error": "파일을 찾을 수 없습니다."
            }), 404

        if not filename.lower().endswith(".xlsx"):
            return jsonify({
                "error": "현재 XLSX 파일만 수정할 수 있습니다."
            }), 400


        # =================================================
        # 숫자 변환
        # =================================================

        def to_number(value):

            if value is None:
                return 0

            if isinstance(value, (int, float)):
                return value

            text = str(value).strip()

            if not text:
                return 0

            text = text.replace(",", "")

            try:

                number = float(text)

                if number.is_integer():
                    return int(number)

                return number

            except (ValueError, TypeError):

                return 0


        # =================================================
        # 기존 엑셀 열기
        # =================================================

        workbook = openpyxl.load_workbook(
            file_path,
            data_only=False
        )

        sheet = workbook.active


        # =================================================
        # 새 엑셀 구조
        #
        # 0 제품명
        # 1 단가
        # 2 입고수량
        # 3 재고
        # 4 판매수량
        # 5 판매금액
        # 6 비고
        # 7 메모
        # =================================================

        clean_rows = []


        # 최대 48행까지만 상품 데이터
        product_rows = rows[:48]


        for row in product_rows:

            if row is None:
                row = []

            if not isinstance(row, (list, tuple)):
                row = []

            row = list(row)


            # 최소 8열 확보
            while len(row) < 8:
                row.append("")


            # =================================================
            # 값 읽기
            # =================================================

            product_name = (
                str(row[0]).strip()
                if row[0] is not None
                else ""
            )


            price = to_number(
                row[1]
            )


            incoming = to_number(
                row[2]
            )


            stock = to_number(
                row[3]
            )


            sales_qty = to_number(
                row[4]
            )


            # =================================================
            # 판매금액 자동 계산
            #
            # 단가 × 판매수량
            # =================================================

            if (
                product_name != "" and
                str(row[1]).strip() != "" and
                str(row[4]).strip() != ""
            ):

                sales_amount = (
                    price * sales_qty
                )
            else:

                sales_amount = ""


            # =================================================
            # 저장용 행
            #
            # 입고수량은 그대로 저장
            # 재고에는 자동 합산하지 않음
            # =================================================

            clean_row = [

                product_name,

                price if str(row[1]).strip() != ""
                else "",

                incoming if str(row[2]).strip() != ""
                else "",

                stock if str(row[3]).strip() != ""
                else "",

                sales_qty if str(row[4]).strip() != ""
                else "",

                sales_amount,

                row[6]
                if row[6] is not None
                else "",

                row[7]
                if row[7] is not None
                else ""

            ]


            clean_rows.append(
                clean_row
            )


        # =================================================
        # 49행
        #
        # A = 합계금액
        # B = 전체 판매금액
        # C = 입금액
        # D = 직접 입력
        # E = 잔액
        # F = 합계금액 - 입금액
        # =================================================

        deposit_amount = 0


        if len(rows) >= 49:

            row49 = rows[48]

            if isinstance(
                row49,
                (list, tuple)
            ):

                if len(row49) > 3:

                    deposit_amount =
                        to_number(
                            row49[3]
                        )


        # =================================================
        # 전체 판매금액
        # =================================================

        total_sales_amount = 0


        for row in clean_rows:

            total_sales_amount += \
                to_number(row[5])


        # =================================================
        # 잔액
        # =================================================

        balance = (
            total_sales_amount -
            deposit_amount
        )


        summary_row = [

            "합계금액",

            total_sales_amount,

            "입금액",

            deposit_amount,

            "잔액",

            balance,

            "",

            ""

        ]


        # =================================================
        # 50행
        #
        # A = 재고금액
        # B = 단가 × 재고 전체 합계
        # =================================================

        total_stock_amount = 0


        for row in clean_rows:

            total_stock_amount += (
                to_number(row[1]) *
                to_number(row[3])
            )


        stock_total_row = [

            "재고금액",

            total_stock_amount,

            "",

            "",

            "",

            "",

            "",

            ""

        ]


        # =================================================
        # 기존 엑셀 전체 삭제
        # =================================================

        if sheet.max_row > 0:

            sheet.delete_rows(
                1,
                sheet.max_row
            )


        # =================================================
        # 1~48행 상품 저장
        # =================================================

        for row in clean_rows:

            sheet.append(
                row
            )


        # =================================================
        # 49행 저장
        # =================================================

        sheet.append(
            summary_row
        )


        # =================================================
        # 50행 저장
        # =================================================

        sheet.append(
            stock_total_row
        )


        # =================================================
        # 숫자 형식
        #
        # 2 단가
        # 3 입고수량
        # 4 재고
        # 5 판매수량
        # 6 판매금액
        # =================================================

        for row_num in range(
            1,
            sheet.max_row + 1
        ):

            for col_num in [
                2,
                3,
                4,
                5,
                6
            ]:

                sheet.cell(
                    row=row_num,
                    column=col_num
                ).number_format = '#,##0'


        # =================================================
        # 저장
        # =================================================

        workbook.save(
            file_path
        )


        return jsonify({
            "success": True
        })


    except Exception as e:

        print(
            "엑셀 저장 오류:",
            e
        )

        return jsonify({
            "error": str(e)
        }), 500

@app.route("/preview/<path:filename>")
def preview(filename):

    if not session.get("admin"):
        return jsonify({"error": "unauthorized"})

    return jsonify(get_preview(filename))

@app.route("/analyze", methods=["POST"])
def analyze():

    try:
 
        clean_old_files()
        order = request.files.get("order_file")
        bank = request.files.get("bank_file")

        order_text = request.form.get("order_text", "").strip()

        ALLOWED_EXTENSIONS = (".xlsx", ".xls")

        # -------------------------
        # 주문 읽기
        # -------------------------
        if order and order.filename:

            if not order.filename.lower().endswith(ALLOWED_EXTENSIONS):
                raise SellerOSError("EH001")

            order_name = secure_filename(order.filename)

            order_path = os.path.join(
                "uploads",
                "orders",
                order_name
            )

            order.save(order_path)
                    
            orders = load_order_excel(order_path)
      

        elif order_text:

            orders = load_order_text(order_text)

        else:

            raise SellerOSError("EH001")

        # -------------------------
        # 입금파일 없는 경우
        # → 주문목록 생성
        # -------------------------
        if not bank or bank.filename == "":

            filename = datetime.now().strftime("%Y%m%d_%H%M%S_주문목록.xlsx")

            output_file = os.path.join(
              "output",
              filename
)

            save_order_list_excel(
                orders,
                output_file
            )

            preview = group_preview(orders)

            session["result_file"] = output_file

            return render_template(
                "result.html",
                total=len(orders),
                preview=preview[:10],
                matched=0,
                unmatched=0,
                no_order=0,
                order_only=True

            )

        # -------------------------
        # 입금파일 검사
        # -------------------------
        if not bank.filename.lower().endswith(ALLOWED_EXTENSIONS):
            raise SellerOSError("EH001")

        bank_name = secure_filename(bank.filename)

        bank_path = os.path.join(
            "uploads",
            "banks",
            bank_name
        )

        bank.save(bank_path)

        # -------------------------
        # 자동 검수
        # -------------------------
        result = run_pipeline(
            bank_file=bank_path,
            orders=orders
        )

        session["result_file"] = result["result_file"]
        session["logen_file"] = result["logen_file"]

        return render_template(
            "result.html",
            total=result["total"],
            matched=result["matched"],
            unmatched=result["unmatched"],
            no_order=result["no_order"],
            preview=result["preview"][:10]

        )

    except SellerOSError as e:

        return render_template(
            "error.html",
            code=e.code,
            message=e.message
        )

    except Exception as e:

        print(e)

        return render_template(
            "error.html",
            code="EH999",
            message="프로그램 처리 중 오류가 발생했습니다."
    )

@app.route("/download/result")
def download_result():

    return send_file(
        session["result_file"],
        as_attachment=True
    )


@app.route("/download/logen")
def download_logen():

    return send_file(
        session["logen_file"],
        as_attachment=True
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
